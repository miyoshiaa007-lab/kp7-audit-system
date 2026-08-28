import streamlit as st
import pandas as pd
import json
import re
import io
import os
import tempfile
import time
import math
from datetime import datetime
from pydantic import BaseModel, Field
from typing import List, Optional, Tuple, Dict, Any

# นำเข้า Google GenAI SDK รุ่นใหม่ล่าสุด
from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ==========================================
# 0. ตั้งค่าหน้าเว็บ และ CSS (ซ่อนช่องว่างล่องหน)
# ==========================================
st.set_page_config(page_title="e-KP7 Audit System", page_icon="🎯", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
    <style>
    /* ซ่อนช่องว่างล่องหนด้านบน */
    .element-container:has(> iframe[title="st.markdown"]) { display: none; }
    .main {background-color: #F8F9FA;}
    h1 {color: #1F4E79; font-weight: 700; margin-top: -1.5rem;}
    .stTabs [data-baseweb="tab-list"] {gap: 20px;}
    .stTabs [data-baseweb="tab"] {padding: 10px 20px; border-radius: 8px 8px 0px 0px;}
    .stMetric {background-color: white; padding: 15px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.05); border-left: 5px solid #1F4E79;}
    </style>
""", unsafe_allow_html=True)

# ฐานในการคำนวณเงินเดือน ตามเกณฑ์ ก.ค.ศ. (ใช้ได้เฉพาะระบบ "อันดับ/ระดับ" แบบ คศ. ปี 2551 เป็นต้นมา
# เอกสารจริงก่อนหน้านั้น (เช่น อาจารย์ 1 ระดับ 3-6 ปี 2537-2547 ใน otepc63.pdf) ไม่มีคำว่า "คศ." เลย
# -> normalize_standing จะคืนค่าว่าง และข้ามการตรวจสอบยอดคำนวณให้โดยอัตโนมัติ ซึ่งถูกต้องแล้ว)
SALARY_BASES = {
    "คศ.5": {"split": 60840, "upper": 68560, "lower": 60830},
    "คศ.4": {"split": 50330, "upper": 59630, "lower": 50320},
    "คศ.3": {"split": 40280, "upper": 49330, "lower": 37200},
    "คศ.2": {"split": 30210, "upper": 35270, "lower": 30200},
    "คศ.1": {"split": 24890, "upper": 29600, "lower": 22780},
    "ครูผู้ช่วย": {"split": 19910, "upper": 22330, "lower": 17480}
}

def normalize_standing(acad_str: str) -> str:
    """
    แก้ไข (เทียบกับไฟล์จริง otepc63.pdf):
    เดิมมีเงื่อนไข `"ครู" in text and "ผู้ช่วย" not in text` ซึ่งจับคำว่า "ครู" แบบ substring
    กว้างเกินไป ทดสอบแล้วพบว่าข้อความอย่าง "ครู ร.ร.บ้านหนองหว้า..." หรือแม้แต่คำว่า
    "พนักงานราชการครู" จะถูกจัดเป็น "คศ.1" ทั้งที่ไม่เกี่ยวกับวิทยฐานะเลย
    เอกสารจริงจากระบบ HRMS จะระบุ คศ.1-5 ไว้ชัดเจนในคอลัมน์ "อันดับ/ระดับ" อยู่แล้วเสมอ
    (เช่น "คศ.2", "คศ.3" ใน otepc63.pdf หน้า 4-13) จึงตัดการเดาจากคำว่า "ครู" ลอยๆ ทิ้งไปเลย
    เพื่อไม่ให้เกิด false positive — กรณีไม่พบคำที่ชัดเจนจะคืนค่าว่าง (ปลอดภัยกว่า เพราะแค่ทำให้
    ข้ามการตรวจสอบยอดคำนวณสำหรับแถวนั้น ไม่ใช่ไปสร้างคำเตือนเท็จ)
    """
    if not acad_str: return ""
    text = str(acad_str).replace(" ", "")
    if "เชี่ยวชาญพิเศษ" in text or "คศ.5" in text: return "คศ.5"
    if "เชี่ยวชาญ" in text or "คศ.4" in text: return "คศ.4"
    if "ชำนาญการพิเศษ" in text or "คศ.3" in text: return "คศ.3"
    if "ชำนาญการ" in text or "คศ.2" in text: return "คศ.2"
    if "ครูผู้ช่วย" in text: return "ครูผู้ช่วย"
    if "คศ.1" in text: return "คศ.1"
    return ""

def extract_percent_value(text: str) -> Optional[float]:
    """
    แก้ไข (เทียบกับไฟล์จริง otepc63.pdf):
    เดิมใช้ regex r'(\\d+\\.\\d{1,2})' จับตัวเลขทศนิยมใดๆ ใน percentage_or_step แล้วตีความเป็น "เปอร์เซ็นต์"
    ทันที แต่ในเอกสารจริงพบว่ามี "ระบบเลื่อนขั้นแบบขั้น" ปนอยู่ เช่น "เลื่อนขั้นเงินเดือน 0.5 ขั้น"
    (ระบบเก่า ก่อนปี ~2559 ที่ใช้กันตลอด ตั้งแต่ 2549-2558 ในไฟล์ตัวอย่าง) ซึ่งเลข "0.5" ที่นี่คือ
    "ครึ่งขั้น" ไม่ใช่ 0.5% — ถ้าเอาไปคูณฐานเงินเดือน/100 ตามสูตรร้อยละ จะได้ยอด "ควรเป็น" ที่ผิดทุกแถว
    ระบบร้อยละของจริง (เช่น "เลื่อนเงินเดือน ร้อยละ 3.03 (ดีเด่น)" ที่พบตั้งแต่ปี 2562 เป็นต้นไปในไฟล์ตัวอย่าง)
    จะมีคำว่า "ร้อยละ" หรือ "%" กำกับเสมอ จึงต้องเช็คคำกำกับก่อน ไม่ใช่จับตัวเลขลอยๆ
    """
    text = str(text or "")
    if "ขั้น" in text and "ร้อยละ" not in text and "%" not in text:
        return None
    match_pct = re.search(r'ร้อยละ\s*(\d+(?:\.\d{1,2})?)', text)
    if not match_pct:
        match_pct = re.search(r'(\d+(?:\.\d{1,2})?)\s*%', text)
    if match_pct:
        return float(match_pct.group(1))
    return None

def calculate_new_salary(old_salary: float, standing_str: str, percent: float) -> Optional[float]:
    standing = normalize_standing(standing_str)
    if not standing or standing not in SALARY_BASES or not old_salary or not percent: return None
    bases = SALARY_BASES[standing]
    base_salary = bases["upper"] if old_salary >= bases["split"] else bases["lower"]
    increment = base_salary * (percent / 100.0)
    increment_rounded = math.ceil(increment / 10.0) * 10 # ปัดเศษขึ้นเป็นหลักสิบ
    return old_salary + increment_rounded

# ==========================================
# 1. โครงสร้างข้อมูล & 2. ทำความสะอาดข้อมูล
# ==========================================
class RecordEntry(BaseModel):
    date_raw: str = Field(default="")
    position_and_workplace: str = Field(default="")
    position_no: Optional[str] = Field(default="")
    academic_standing: Optional[str] = Field(default="")   # มาจากคอลัมน์ "อันดับ/ระดับ" เป็นหลัก (ดูคำอธิบายใน prompt)
    salary: Optional[float] = Field(default=0.0)
    order_ref: Optional[str] = Field(default="")
    percentage_or_step: Optional[str] = Field(default="")
    reason_for_update: Optional[str] = Field(default="")

class KP7ExtractionResult(BaseModel):
    records: list[RecordEntry] = Field(default=[])

THAI_DIGITS = str.maketrans("๐๑๒๓๔๕๖๗๘๙", "0123456789")
THAI_MONTHS = {"ม.ค.": 1, "มค": 1, "มกราคม": 1, "ก.พ.": 2, "กพ": 2, "กุมภาพันธ์": 2, "มี.ค.": 3, "มีค": 3, "มีนาคม": 3, "เม.ย.": 4, "เมย": 4, "เมษายน": 4, "พ.ค.": 5, "พค": 5, "พฤษภาคม": 5, "มิ.ย.": 6, "มิย": 6, "มิถุนายน": 6, "ก.ค.": 7, "กค": 7, "กรกฎาคม": 7, "ส.ค.": 8, "สค": 8, "สิงหาคม": 8, "ก.ย.": 9, "กย": 9, "กันยายน": 9, "ต.ค.": 10, "ตค": 10, "ตุลาคม": 10, "พ.ย.": 11, "พย": 11, "พฤศจิกายน": 11, "ธ.ค.": 12, "ธค": 12, "ธันวาคม": 12}
MONTH_LABEL = {1: "ม.ค.", 2: "ก.พ.", 3: "มี.ค.", 4: "เม.ย.", 5: "พ.ค.", 6: "มิ.ย.", 7: "ก.ค.", 8: "ส.ค.", 9: "ก.ย.", 10: "ต.ค.", 11: "พ.ย.", 12: "ธ.ค."}

def clean_json_string(raw_text: str) -> str:
    if not raw_text: return "{}" # ดักจับกรณี AI ไม่ส่งข้อความกลับมา
    text = str(raw_text).strip()
    if text.startswith("```"):
        lines = text.splitlines()
        if len(lines) > 2 and lines[-1].strip().startswith("```"): text = "\n".join(lines[1:-1])
        elif len(lines) > 1: text = "\n".join(lines[1:])
    return text.strip()

def sanitize_salary(sal_val: Any, debug_log: Optional[list] = None) -> float:
    if sal_val is None: return 0.0
    s_str = str(sal_val).translate(THAI_DIGITS).replace(",", "").replace(" ", "").replace("บาท", "")
    match = re.search(r"(\d+(\.\d+)?)", s_str)
    if match:
        val = float(match.group(1))
        if val > 0 and val < 5000:
            fixed = val * 10
            # บันทึกไว้ให้ผู้ตรวจสอบเห็น แทนที่จะแก้ค่าแบบเงียบๆ (สำคัญสำหรับงาน audit ราชการ)
            if debug_log is not None:
                debug_log.append(f"ปรับเงินเดือนอัตโนมัติ: อ่านได้ {val:,.0f} -> ใช้ {fixed:,.0f} บาท (ค่าที่อ่านได้ต่ำกว่า 5,000 บาท ต้องสงสัยว่า OCR ตกหลัก)")
            return fixed
        return val
    return 0.0

def normalize_thai_date(date_str: str) -> Tuple[str, int]:
    if not date_str: return "-", 0
    clean_str = str(date_str).translate(THAI_DIGITS).replace(" ", "")
    pattern = r"(\d{1,2})([ก-๙\.]+)(\d{2,4})"
    match = re.search(pattern, clean_str)
    if not match: return str(date_str).strip(), 0
    day = int(match.group(1))
    month_raw = match.group(2)
    year_raw = int(match.group(3))
    year = 2500 + year_raw if year_raw < 100 else year_raw
    month = next((m_val for m_key, m_val in THAI_MONTHS.items() if m_key in month_raw), 0)
    if month == 0: return f"{day} {month_raw} {year}", (year * 10000) + day
    return f"{day} {MONTH_LABEL[month]} {year}", (year * 10000) + (month * 100) + day

def identify_update_reason(text: str) -> str:
    if not text: return 'เลื่อนปกติ'
    text_str = str(text)
    if re.search(r'แก้ไข', text_str): return 'แก้ไขคำสั่ง'
    elif re.search(r'พ\.ร\.บ\.|พรบ|พ\.ร\.ฎ\.|ปรับตาม', text_str): return 'ปรับตาม พ.ร.บ./พ.ร.ฎ.'
    elif re.search(r'ชดเชย|ปรับอัตรา', text_str): return 'ปรับชดเชยมติ ครม.'
    else: return 'เลื่อนปกติ'

# ==========================================
# 3. VLM Data Extractor (SDK API Check)
# ==========================================
def extract_pdf_records_precise(pdf_bytes: bytes, api_key: str, model_name: str, hint: str, debug_log: Optional[list] = None) -> List[Dict[str, Any]]:
    client = genai.Client(api_key=api_key)
    # ปรับปรุง prompt ตามโครงสร้างคอลัมน์จริงที่เจอในเอกสาร HRMS (otepc63.pdf) และเล่มเขียนมือ (ก.ค.ศ.16):
    # วัน เดือน ปี | ตำแหน่ง/หน่วยงานการศึกษา/ส่วนราชการ | ตำแหน่งเลขที่ | วิทยฐานะ | ตำแหน่งประเภท |
    # อันดับ/ระดับ | อัตราเงินเดือน/ค่าตอบแทนพิเศษ | เงินวิทยฐานะ/เงินประจำตำแหน่ง/เงินค่าตอบแทน |
    # เงินเพิ่มพิเศษอื่นๆ | เอกสารอ้างอิง
    prompt = f"""คุณคือผู้เชี่ยวชาญการตรวจสอบทะเบียนประวัติ สพป.มหาสารคาม เขต 2 เอกสารนี้คือ: {hint}
    เอกสารนี้เป็นตารางที่มีคอลัมน์ (ชื่ออาจสะกดต่างกันเล็กน้อยระหว่างเล่มเขียนมือกับระบบอิเล็กทรอนิกส์ แต่ความหมายตรงกัน):
    วัน เดือน ปี | ตำแหน่ง/หน่วยงานการศึกษา/ส่วนราชการ | ตำแหน่งเลขที่ | วิทยฐานะ | ตำแหน่งประเภท |
    อันดับ/ระดับ | อัตราเงินเดือน/ค่าตอบแทนพิเศษ | เงินวิทยฐานะ/เงินประจำตำแหน่ง | เอกสารอ้างอิง

    กฎการสกัดข้อมูล (สกัดทุกแถว ห้ามข้าม แม้จะดูซ้ำกับแถวก่อนหน้า):
    1. date_raw = คอลัมน์ "วัน เดือน ปี"
    2. position_and_workplace = คอลัมน์ "ตำแหน่ง/หน่วยงานการศึกษา/ส่วนราชการ" (ข้อความเต็ม)
    3. position_no = คอลัมน์ "ตำแหน่งเลขที่"
    4. academic_standing = ให้ใช้ค่าจากคอลัมน์ "อันดับ/ระดับ" เป็นหลัก (เช่น "คศ.1"-"คศ.5") เพราะเป็นค่าที่
       ผูกกับฐานเงินเดือนจริง **ห้ามใช้คอลัมน์ "วิทยฐานะ" แทน** เนื่องจากบางช่วงเวลาสองคอลัมน์นี้ไม่ตรงกัน
       (เช่น วิทยฐานะยังโชว์ "ชำนาญการ" ทั้งที่อันดับ/ระดับขยับเป็น "คศ.3" แล้ว ให้ยึดอันดับ/ระดับเป็นหลักเสมอ)
    5. salary = ตัวเลขในคอลัมน์ "อัตราเงินเดือน/ค่าตอบแทนพิเศษ"
    6. order_ref = คอลัมน์ "เอกสารอ้างอิง" (คัดลอกให้ครบ อย่าตัดทอน เพราะบางวันที่ซ้ำกันหลายแถว เอกสารอ้างอิงคือ
       สิ่งเดียวที่แยกแถวออกจากกันได้)
    7. percentage_or_step = ถ้าข้อความระบุ "ร้อยละ X" หรือ "X%" ให้เก็บเป็น "X%" ถ้าระบุ "X ขั้น" (เช่น "0.5 ขั้น",
       "1 ขั้น") ให้เก็บเป็น "X ขั้น" **ห้ามปนกัน** เพราะเป็นคนละระบบคำนวณ (ขั้น = ระบบเก่าก่อน พ.ศ. 2559,
       ร้อยละ = ระบบใหม่)
    8. reason_for_update = สรุปเหตุผลถ้าระบุชัด (แก้ไขคำสั่ง / ปรับตาม พ.ร.บ. หรือ พ.ร.ฎ. / ปรับชดเชยมติ ครม.)

    ข้อควรระวังสำคัญ (พบได้ในทุกปีที่มีการปรับกฎหมายเงินเดือนข้าราชการ ไม่ใช่เฉพาะปีล่าสุด):
    - จะมีแถว "ปรับย้อนหลัง/ปรับชดเชย/แก้ไขคำสั่ง" แทรกเข้ามาโดยมี "วันที่" ซ้ำกับแถวเดิมที่เคยบันทึกไปแล้ว
      แต่เงินเดือนหรือเอกสารอ้างอิงต่างกัน (พบได้ตั้งแต่ปี 2552, 2554, 2558 ไปจนถึงปีล่าสุด) ให้บันทึกทุกแถว
      แยกกันตามจริง ห้ามรวมหรือข้ามแถวที่ดูเหมือนซ้ำ
    - บางแถวในระบบอิเล็กทรอนิกส์จะมีคู่แถวที่ "วันที่และเงินเดือนเดียวกัน" แต่คำอธิบายตำแหน่งต่างกัน (แถวสรุปสั้นๆ
      กับแถวอธิบายเต็ม) ให้บันทึกทั้งสองแถวแยกกันตามที่เห็นจริง"""

    temp_pdf_path = ""
    uploaded_file = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
            temp_pdf.write(pdf_bytes)
            temp_pdf_path = temp_pdf.name

        # ส่งไฟล์ขึ้น Google GenAI Server
        uploaded_file = client.files.upload(file=temp_pdf_path)

        # รอสถานะไฟล์จนกว่าจะพร้อม (Poling ป้องกัน Error 400/500)
        waited = 0
        file_info = client.files.get(name=uploaded_file.name)
        while str(getattr(file_info, "state", "")).upper() in ["PROCESSING", "STATE_PROCESSING"] and waited < 60:
            time.sleep(3)
            waited += 3
            file_info = client.files.get(name=uploaded_file.name)
        if str(getattr(file_info, "state", "")).upper() in ["PROCESSING", "STATE_PROCESSING"]:
            raise TimeoutError("ไฟล์ยังประมวลผลไม่เสร็จภายใน 60 วินาที กรุณาลองใหม่อีกครั้ง")

        # สั่งประมวลผลพร้อมจัด Format JSON
        response = client.models.generate_content(
            model=model_name, contents=[uploaded_file, prompt],
            config=types.GenerateContentConfig(response_mime_type="application/json", response_schema=KP7ExtractionResult, temperature=0.0)
        )
        cleaned_str = clean_json_string(response.text)
    finally:
        # เคลียร์ข้อมูลขยะออกจากระบบ
        if uploaded_file:
            try: client.files.delete(name=uploaded_file.name)
            except Exception as e:
                if debug_log is not None: debug_log.append(f"ลบไฟล์ชั่วคราวบน Google server ไม่สำเร็จ: {e}")
        if temp_pdf_path and os.path.exists(temp_pdf_path):
            try: os.remove(temp_pdf_path)
            except Exception as e:
                if debug_log is not None: debug_log.append(f"ลบไฟล์ temp ในเครื่องไม่สำเร็จ: {e}")

    try:
        data = json.loads(cleaned_str).get("records", [])
    except Exception as e:
        data = []
        if debug_log is not None:
            debug_log.append(f"แปลง JSON จากคำตอบ AI ไม่สำเร็จ ({e}) — คำตอบดิบ: {cleaned_str[:500]}")

    extracted_rows = []
    for idx, r in enumerate(data):
        r["normalized_date"], r["sort_key"] = normalize_thai_date(r.get("date_raw", ""))
        r["salary"] = sanitize_salary(r.get("salary", 0), debug_log=debug_log)
        r["original_index"] = idx
        if not r.get("reason_for_update"):
            r["reason_for_update"] = identify_update_reason(str(r.get("position_and_workplace", "")) + " " + str(r.get("reason_for_update", "")))
        extracted_rows.append(r)
    return extracted_rows

# ==========================================
# 4. Smart Reconciliation (แกนหลักประมวลผล)
# ==========================================
def format_milestone_desc(record: dict) -> str:
    # ใช้ or '' เพื่อเปลี่ยนค่า None ให้เป็นข้อความเปล่าอัตโนมัติ
    pos = str(record.get('position_and_workplace') or '')
    pos_no = str(record.get('position_no') or '')
    acad = str(record.get('academic_standing') or '')
    pct = str(record.get('percentage_or_step') or '')
    order = str(record.get('order_ref') or '')
    salary = float(record.get('salary') or 0.0)

    desc = f"{pos} "
    tags_raw = [f"เลข:{pos_no}", acad, f"เลื่อน:{pct}"]
    # ตัดช่องว่างอย่างปลอดภัย
    tags = [t.strip() for t in tags_raw if t.strip() and t.strip() not in ["เลข:", "เลื่อน:"]]
    tag_str = f"[{' | '.join(tags)}]" if tags else ""

    return f"{desc.strip()} {tag_str} (เงินเดือน {salary:,.0f} บ.) [{order}]"

def find_best_match_pairs(records_a: List[Dict[str, Any]], records_b: List[Dict[str, Any]]) -> Dict[int, int]:
    """
    แก้ไข (ยืนยันจากไฟล์จริง otepc63.pdf): พบเงินเดือนซ้ำกันบ่อยมากในเอกสารจริง เช่น หน้า 5-9
    มีหลายแถวที่ "วันที่และเงินเดือนเดียวกัน" (คู่แถวสรุป/แถวอธิบายเต็มของ HRMS) และยังพบแถวที่
    "วันที่เดียวกันแต่เงินเดือนต่างกัน" ถึง 4 แถวติดกัน (กรณีแก้ไขคำสั่งซ้อนหลายรอบในวันเดียว เช่น
    1 เม.ย. 2554 มี 4 แถว: 25,190 / 26,260 / 27,580 / 26,450 บาท)

    โค้ดเดิมวนหา records_b ที่เงินเดือนตรงกันตัวแรกที่ยังไม่ถูกใช้ แล้ว "เขียนทับ" matched_b_idx ไปเรื่อยๆ
    จนสุดลูปถ้าไม่เจอวันที่ตรงเป๊ะ ผลคือได้ "แถวสุดท้าย" ที่เงินเดือนตรง ไม่ใช่แถวที่วันที่ใกล้เคียงที่สุด
    ทำให้จับคู่ผิดแถว แล้ว mark ว่า "ใช้แล้ว" ไปทำให้แถวที่ควรจะจับคู่ได้จริงในรอบถัดไปหาคู่ไม่เจอ
    กลายเป็นแจ้งเตือน "ขาดในเขียนมือ/ขาดใน HRMS" ที่ผิดพลาดต่อเนื่องกันเป็นทอดๆ

    ฟังก์ชันนี้แก้โดยรวบรวมคู่ที่เป็นไปได้ทั้งหมดก่อน (เงินเดือนตรงกัน) แล้วจับคู่แบบ greedy โดยเรียง
    ตามผลต่างวันที่ (sort_key) จากน้อยไปมาก เพื่อให้คู่ที่ "วันที่ใกล้กันที่สุด" ได้จับคู่กันก่อนเสมอ
    """
    candidates = []
    for i, r_a in enumerate(records_a):
        for j, r_b in enumerate(records_b):
            if abs(r_a["salary"] - r_b["salary"]) < 1.0:
                if r_a["sort_key"] > 0 and r_b["sort_key"] > 0:
                    date_diff = abs(r_a["sort_key"] - r_b["sort_key"])
                else:
                    date_diff = 10**9  # ไม่มีวันที่ให้เทียบ ให้ความสำคัญต่ำสุด
                candidates.append((date_diff, i, j))
    candidates.sort(key=lambda x: x[0])

    match_a_to_b, used_a, used_b = {}, set(), set()
    for _, i, j in candidates:
        if i in used_a or j in used_b:
            continue
        match_a_to_b[i] = j
        used_a.add(i)
        used_b.add(j)
    return match_a_to_b

def run_two_way_reconciliation(records_a: List[Dict[str, Any]], records_b: List[Dict[str, Any]]):
    inversions_b = []
    for i in range(1, len(records_b)):
        prev_b, curr_b = records_b[i-1], records_b[i]
        if curr_b["sort_key"] > 0 and prev_b["sort_key"] > 0 and curr_b["sort_key"] < prev_b["sort_key"]:
            inversions_b.append({"msg": f"บรรทัดที่ {curr_b['original_index']+1} ({curr_b.get('date_raw')}) จดย้อนหลังสลับกับบรรทัดก่อนหน้า ({prev_b.get('date_raw')})"})

    for i in range(1, len(records_a)):
        records_a[i]["is_transfer"] = bool(records_a[i].get("position_no") and records_a[i-1].get("position_no") and records_a[i].get("position_no") != records_a[i-1].get("position_no"))
        records_a[i]["is_promotion"] = bool(records_a[i].get("academic_standing") and records_a[i-1].get("academic_standing") and records_a[i].get("academic_standing") != records_a[i-1].get("academic_standing"))

    match_a_to_b = find_best_match_pairs(records_a, records_b)
    used_b_indices = set(match_a_to_b.values())

    matched_rows, stats, prev_salary_a, seq = [], {"perfect_match": 0, "missing_in_manual": 0, "missing_in_hrms": 0}, 0.0, 0

    for idx_a, r_a in enumerate(records_a):
        flag_msg = ""
        if r_a.get("is_transfer"): flag_msg += "🚩 เปลี่ยนเลขตำแหน่ง "
        if r_a.get("is_promotion"): flag_msg += "🌟 เลื่อนวิทยฐานะ "
        if r_a.get('reason_for_update') in ['แก้ไขคำสั่ง', 'ปรับตาม พ.ร.บ./พ.ร.ฎ.']: flag_msg += f"✏️ {r_a.get('reason_for_update')} "
        if "ชดเชย" in str(r_a.get("position_and_workplace", "")) or ("1 พ.ค." in r_a["normalized_date"]): flag_msg += "💰 ปรับชดเชย ครม. "

        # แก้ไข (ยืนยันจากไฟล์จริง): ใช้ extract_percent_value แทน regex เดิมที่จับ "0.5 ขั้น" ผิดเป็น 0.5%
        percent_val = extract_percent_value(r_a.get('percentage_or_step', ''))

        if percent_val and prev_salary_a > 0 and r_a.get("reason_for_update") == "เลื่อนปกติ":
            calc_sal = calculate_new_salary(prev_salary_a, r_a.get("academic_standing"), percent_val)
            if calc_sal:
                if abs(calc_sal - r_a["salary"]) <= 20: flag_msg += "[🧮 ยอดคำนวณ: เป๊ะ]"
                else: flag_msg += f"[⚠️ ควรเป็น {calc_sal:,.0f} บ.]"

        prev_salary_a = r_a["salary"]
        matched_b_idx = match_a_to_b.get(idx_a)

        if matched_b_idx is not None:
            r_b = records_b[matched_b_idx]
            status = "⚠️ เงินเดือนตรง/วันที่คลาดเคลื่อน" if r_a["normalized_date"] != r_b["normalized_date"] else "✅ ตรงกันสมบูรณ์"
            stats["perfect_match"] += 1
            matched_rows.append({"_seq": seq, "วัน เดือน ปี": r_a["normalized_date"], "เงินเดือน": f"{r_a['salary']:,.0f}", "HRMS (อิเล็กทรอนิกส์)": format_milestone_desc(r_a), "ก.ค.ศ.16 (เขียนมือ)": format_milestone_desc(r_b), "สถานะการตรวจสอบ": f"{status} {flag_msg}".strip(), "สิ่งที่ต้องแก้": "-"})
        else:
            stats["missing_in_manual"] += 1
            matched_rows.append({"_seq": seq, "วัน เดือน ปี": r_a["normalized_date"], "เงินเดือน": f"{r_a['salary']:,.0f}", "HRMS (อิเล็กทรอนิกส์)": format_milestone_desc(r_a), "ก.ค.ศ.16 (เขียนมือ)": "-", "สถานะการตรวจสอบ": f"❌ ขาดในเขียนมือ {flag_msg}".strip(), "สิ่งที่ต้องแก้": "เพิ่มลงสมุด ก.ค.ศ.16"})
        seq += 1

    for idx_b, r_b in enumerate(records_b):
        if idx_b not in used_b_indices:
            stats["missing_in_hrms"] += 1
            matched_rows.append({"_seq": seq, "วัน เดือน ปี": r_b["normalized_date"], "เงินเดือน": f"{r_b['salary']:,.0f}", "HRMS (อิเล็กทรอนิกส์)": "-", "ก.ค.ศ.16 (เขียนมือ)": format_milestone_desc(r_b), "สถานะการตรวจสอบ": "❌ ขาดใน HRMS", "สิ่งที่ต้องแก้": "คีย์เข้าระบบ e-KP7"})
            seq += 1

    def get_sort_val(row):
        try: pts = row["วัน เดือน ปี"].split(); return (int(pts[2]) * 10000 + THAI_MONTHS.get(pts[1], 0) * 100 + int(pts[0]), row["_seq"])
        except: return (99999999, row["_seq"])
    result = sorted(matched_rows, key=get_sort_val)
    for row in result: row.pop("_seq", None)
    return result, stats, inversions_b

# การระบายสีตาราง Pandas DataFrame เพื่อ UI ที่สวยงาม
def highlight_status(row):
    status = str(row['สถานะการตรวจสอบ'])
    if '✅' in status: return ['background-color: #E6F4EA; color: #137333'] * len(row)
    elif '❌' in status: return ['background-color: #FCE8E6; color: #A50E0E'] * len(row)
    elif '⚠️' in status: return ['background-color: #FEF7E0; color: #B06000'] * len(row)
    return [''] * len(row)

def generate_audit_excel(table_rows) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "ผลตรวจสอบ"
    ws.append(["ลำดับ", "วัน เดือน ปี", "เงินเดือน", "ก.พ.7 (HRMS)", "ก.ค.ศ.16 (เขียนมือ)", "สถานะ", "การแก้ไข"])
    for cell in ws[1]: cell.fill = PatternFill(start_color="1F4E79", fill_type="solid"); cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center")
    for idx, r in enumerate(table_rows, 1): ws.append([idx, r["วัน เดือน ปี"], r["เงินเดือน"], r["HRMS (อิเล็กทรอนิกส์)"], r["ก.ค.ศ.16 (เขียนมือ)"], r["สถานะการตรวจสอบ"], r["สิ่งที่ต้องแก้"]])
    ws.column_dimensions['D'].width = ws.column_dimensions['E'].width = 45; ws.column_dimensions['F'].width = 35
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ==========================================
# 5. UI แบบใหม่ (Modern Layout)
# ==========================================
st.title("🎯 ระบบประมวลผลเทียบเคียง ก.พ.7 / ก.ค.ศ.16")
st.caption("พัฒนาเพื่อ สพป.มหาสารคาม เขต 2 (ขับด้วย Gemini GenAI)")

with st.sidebar:
    st.header("⚙️ การตั้งค่า AI (Hybrid Mode)")

    # ดึง API Key จากหลังบ้าน (สพป.) อัตโนมัติ (ถ้ามี)
    if "GEMINI_API_KEY" in st.secrets:
        api_key_input = st.secrets["GEMINI_API_KEY"]
        st.success("✅ เชื่อมต่อระบบ API ของ สพป. เรียบร้อยแล้ว (พร้อมแชร์ลิงก์ให้ทีมงานใช้ได้เลย)")
    else:
        api_key_input = st.text_input("🔑 ใส่ Google Gemini API Key:", type="password")

    # แยกรุ่น AI ทำงานตามความเหมาะสมของไฟล์ (ตรวจสอบชื่อรุ่นให้ตรงกับที่เปิดให้ใช้งานจริงก่อนใช้งานจริงเสมอ)
    st.subheader("แยกประมวลผล (ความเร็ว+ความแม่นยำ)")
    model_hrms = st.selectbox("🤖 อ่านแฟ้มระบบ ก.พ.7 (เน้นไว):", [
        "gemini-3.6-flash",
        "gemini-3.7-flash"
    ], index=0)
    # แก้ไข (ยืนยันจากการเรียก API จริง): "gemini-3.6-pro" ไม่มีอยู่จริง ได้ 404 NOT_FOUND
    # โมเดลระดับ "pro" ที่มีจริงในบัญชีนี้คือ gemini-3.1-pro-preview / gemini-pro-latest
    # (ทดสอบ gemini-3.1-pro-preview กับไฟล์เขียนมือจริงแล้วสกัดข้อมูลได้ปกติ)
    model_man = st.selectbox("🧠 แกะลายมือ ก.ค.ศ.16 (เน้นแม่น):", [
        "gemini-3.1-pro-preview",  # <--- ตัวท็อปเรื่องการแกะลายมือและวิเคราะห์ตาราง (ยืนยันเรียกได้จริง)
        "gemini-pro-latest",
        "gemini-3.6-flash",
        "gemini-3.7-flash"
    ], index=0)
    st.info("💡 นำเข้าข้อมูลการประเมินเงินเดือนตามฐานการคำนวณของ ก.ค.ศ. เรียบร้อยแล้ว")

tab1, tab2, tab3 = st.tabs(["📂 1. อัปโหลดเอกสาร", "📊 2. ผลการตรวจสอบ", "📥 3. สรุป & ดาวน์โหลด"])

with tab1:
    col1, col2 = st.columns(2)
    with col1: file_hrms = st.file_uploader("📄 ก.พ.7 อิเล็กทรอนิกส์ (จากระบบ)", type=["pdf"])
    with col2: file_manual = st.file_uploader("✍️ ก.ค.ศ.16 (เล่มเขียนมือ)", type=["pdf"])

    start_btn = st.button("🚀 เริ่มการตรวจสอบ (Start Audit)", type="primary", use_container_width=True)

if start_btn:
    if not api_key_input or not file_hrms or not file_manual:
        st.error("⚠️ กรุณากรอกข้อมูล/อัปโหลดไฟล์ให้ครบทั้ง 2 ช่องครับ")
    else:
        with st.spinner('⏳ AI กำลังสกัดและวิเคราะห์ข้อมูล... (อาจใช้เวลา 1-2 นาที)'):
            debug_log = []
            try:
                status_box = st.empty()
                status_box.info(f"📄 1/2 กำลังสกัดข้อมูลแฟ้มประวัติจากระบบ (ใช้ {model_hrms})...")
                rec_hrms = extract_pdf_records_precise(file_hrms.read(), api_key_input, model_hrms, "ก.พ.7 อิเล็กทรอนิกส์", debug_log=debug_log)

                status_box.info(f"✍️ 2/2 กำลังแกะลายมือเล่มประวัติ ก.ค.ศ.16 (ใช้ {model_man})...")
                rec_man = extract_pdf_records_precise(file_manual.read(), api_key_input, model_man, "ก.ค.ศ.16 เขียนมือ", debug_log=debug_log)

                status_box.info("⚖️ กำลังเทียบเคียงข้อมูล ดักจับสลับลำดับ และคำนวณฐานเงินเดือน...")
                comp_results, stats, inversions = run_two_way_reconciliation(rec_hrms, rec_man)

                st.session_state['results'] = comp_results
                st.session_state['stats'] = stats
                st.session_state['inversions'] = inversions
                st.session_state['debug_log'] = debug_log
                st.session_state['run_success'] = True
                status_box.success("✅ ตรวจสอบเสร็จสมบูรณ์! เชิญดูผลลัพธ์ที่แท็บ '2. ผลการตรวจสอบ' ครับ")
            except Exception as e:
                st.session_state['debug_log'] = debug_log
                st.error(f"❌ ระบบขัดข้อง: {str(e)}")

# การจัดการหน้าจอว่างเปล่า (Empty State) เมื่อยังไม่รันข้อมูล
with tab2:
    if 'run_success' in st.session_state and st.session_state['run_success']:
        st.subheader("ตารางเปรียบเทียบข้อมูล (Smart Reconciliation Table)")
        df_tab2 = pd.DataFrame(st.session_state['results'])
        if not df_tab2.empty:
            styled_df = df_tab2.style.apply(highlight_status, axis=1)
            st.dataframe(styled_df, use_container_width=True, height=600)
        else:
            st.warning("⚠️ ไม่พบข้อมูลที่สกัดได้จากไฟล์ (AI อาจจะอ่านไม่ออก หรือไฟล์ถูกเข้ารหัสไว้)")

        if st.session_state['inversions']:
            st.error("⚠️ **พบการจดย้อนหลัง (วันที่สลับลำดับ):**")
            for inv in st.session_state['inversions']: st.write(f"- {inv['msg']}")

        if st.session_state.get('debug_log'):
            with st.expander(f"🛠️ รายละเอียดการปรับ/ข้อผิดพลาดระหว่างสกัดข้อมูล ({len(st.session_state['debug_log'])} รายการ)"):
                for line in st.session_state['debug_log']: st.write(f"- {line}")
    else:
        st.info("👈 กรุณาอัปโหลดไฟล์และกดปุ่ม '🚀 เริ่มการตรวจสอบ' ที่แท็บแรกก่อนครับ")

with tab3:
    if 'run_success' in st.session_state and st.session_state['run_success']:
        st.subheader("ภาพรวมการประมวลผล")
        m1, m2, m3 = st.columns(3)
        m1.metric("✅ ตรงกันสมบูรณ์", f"{st.session_state['stats']['perfect_match']} รายการ")
        m2.metric("❌ ขาดในเล่มเขียนมือ", f"{st.session_state['stats']['missing_in_manual']} รายการ")
        m3.metric("❌ ขาดในระบบ HRMS", f"{st.session_state['stats']['missing_in_hrms']} รายการ")

        df_tab3 = pd.DataFrame(st.session_state['results'])
        if not df_tab3.empty:
            excel_buf = generate_audit_excel(st.session_state['results'])
            st.download_button("📥 ดาวน์โหลดรายงานผล (Excel)", data=excel_buf, file_name=f"Audit_Report_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
    else:
        st.info("👈 ข้อมูลรายงานสรุปจะแสดงขึ้นที่นี่ หลังจากทำการตรวจสอบเสร็จสิ้นครับ")
